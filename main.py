from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import cv2
import numpy as np
import time
import base64


# 加载 Excel 数据
def load_excel_data(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=5, values_only=True):  # 从 E2 开始读取
        if row[0] is None:  # 如果遇到空值，停止读取
            break
        data.append(row[0])
    return data

# 使用 OpenCV 计算滑块目标位置
def find_target_position(background_path, slider_path):
    background = cv2.imread(background_path, 0)  # 灰度模式加载背景图片
    slider = cv2.imread(slider_path, 0)         # 灰度模式加载滑块图片
    result = cv2.matchTemplate(background, slider, cv2.TM_CCOEFF_NORMED)
    _, _, _, max_loc = cv2.minMaxLoc(result)
    return max_loc[0]

# 模拟滑动验证
def solve_slider_challenge(url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(url)
        page.click('.verifyBtn')
        
        # 假设图片保存逻辑保持不变
        slider_image = page.wait_for_selector('#small_img')
        img_src = slider_image.get_attribute('src')
        if img_src.startswith("data:image"):
            base64_data = img_src.split(",")[1]
            image_data = base64.b64decode(base64_data)
            with open('slider.png', "wb") as f:
                f.write(image_data)
        
        background_image = page.wait_for_selector('#cpc_img')
        img_src = background_image.get_attribute('src')
        if img_src.startswith("data:image"):
            base64_data = img_src.split(",")[1]
            image_data = base64.b64decode(base64_data)
            with open('background.png', "wb") as f:
                f.write(image_data)

        # 计算滑块目标位置
        target_x = find_target_position('background.png', 'slider.png')
        
        # 获取滑块元素
        slider = page.wait_for_selector('.drag-box.slideTip .move-img')
        slider_box = slider.bounding_box()
        
        if not slider_box:
            print("未能找到滑块元素！")
            browser.close()
            return

        # 计算起始和目标坐标
        start_x = slider_box['x'] + slider_box['width'] / 2
        start_y = slider_box['y'] + slider_box['height'] / 2
        end_x = slider_box['x'] + target_x
        end_y = slider_box['y'] + slider_box['height'] / 2

        # 记录滑动总距离
        total_slide_distance = abs(end_x - start_x)
        
        # 模拟拖动并记录路径
        steps = 50
        positions = [(start_x, start_y)]  # 记录每一步的坐标
        page.mouse.move(start_x, start_y)
        page.mouse.down()
        
        for i in range(steps):
            current_x = start_x + (end_x - start_x) * (i / steps) + np.random.randint(-2, 3)
            current_y = start_y + np.random.randint(-2, 3)
            page.mouse.move(current_x, current_y)
            positions.append((current_x, current_y))
        
        page.mouse.move(end_x, end_y)
        page.mouse.up()
        positions.append((end_x, end_y))
        
        # 计算模拟移动的总距离
        simulated_distance = 0
        for i in range(1, len(positions)):
            x1, y1 = positions[i-1]
            x2, y2 = positions[i]
            step_distance = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            simulated_distance += step_distance

        print(f"滑动总距离: {total_slide_distance:.2f} 像素")
        print(f"模拟移动的距离: {simulated_distance:.2f} 像素")

        time.sleep(2)
        page.screenshot(path='after_slider.png')
        print("滑块验证完成！")
        browser.close()

# 主程序
if __name__ == "__main__":
    # 加载 Excel 中的 URL
    data = load_excel_data('data.xlsx')
    if not data:
        print("未读取到有效数据！")
    else:
        url = data[0]
        solve_slider_challenge(url)