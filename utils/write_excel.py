from openpyxl import load_workbook


def write_excel(file_path, row, column, data):
    wb = load_workbook(file_path)
    sheet = wb.active
    sheet.cell(row=row, column=column, value=data)
    try:
        wb.save(file_path)
        print(f"第 {row} 行第 {column} 列写入数据：{data}")
    except:
        print("excel写入失败, 可能文件正在被占用")
        input("按任意键退出")
        exit(0)