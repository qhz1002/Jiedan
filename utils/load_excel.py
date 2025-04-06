from openpyxl import load_workbook


# 加载 Excel 数据
def load_excel_urls(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    urls = []
    maxBuyCounts = []
    indexes = []
    productStatuses = []

    for index, row in enumerate(
        sheet.iter_rows(min_row=2, min_col=5, max_col=8, values_only=True), start=2
    ):
        url = row[0]  # E 列 URL
        max_buy = row[1]  # F 列 数量
        product_status = row[3]  # 状态

        if url is None:  # 如果 URL 为空，停止读取
            break

        if product_status == None:  # 如果 H 列状态为空，判断为未处理，添加到爬取列表
            try:
                maxBuyCount = str(max_buy)  # 转换为字符串
                # 添加到列表
                urls.append(url)
                maxBuyCounts.append(maxBuyCount)
                indexes.append(index)
            except ValueError:
                print(f"行 {index}: 无法将 '{max_buy}' 转换为数字，跳过该行")
        else:
            productStatuses.append(None)  # 如果为空，添加 None 占位

    if urls == []:
        print("没有需要爬取的 URL (H列状态已写满)")
        exit(0)

    return urls, maxBuyCounts, indexes
